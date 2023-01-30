import requests
from datetime import datetime, date
import openpyxl
import sys

## TODO import sys, use arguments


url = 'https://api.met.no/weatherapi/locationforecast/2.0/compact?lat=59.63333&lon=11.11667'

headers = {
    'User-Agent': 'My Agent 1.0',
    'From': 'morten.stavik.eggen@gmail.com'
}

response = requests.get(url, headers=headers)
print(response)

data = response.json()

print(data["properties"]["timeseries"][0]["data"]["instant"]["details"]["air_temperature"])

wb = openpyxl.load_workbook('G:\My Drive\Personlige Prosjekt\Været\Strømanalyse2023-1.xlsx')
ws = wb['Ute Temp']
cell_row = datetime.now().timetuple().tm_yday-23
today = date.today()

place = "A"

ws[place+str(cell_row)] = today

daytime = sys.argv[1]

match daytime:
    case "morning":
        place = "B"
    case "mid":
        place = "C"
    case "noon":
        place = "D"

ws[place+str(cell_row)] = data["properties"]["timeseries"][0]["data"]["instant"]["details"]["air_temperature"]

wb.save("G:\My Drive\Personlige Prosjekt\Været\Strømanalyse2023-1.xlsx")